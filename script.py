import json

import time
import mitmproxy.http
from openpyxl import Workbook, load_workbook


def request(flow: mitmproxy.http.HTTPFlow):
    pass


def response(flow: mitmproxy.http.HTTPFlow):
    if 'profile_ext' in flow.request.url:
        result = json.loads(flow.response.get_text())
        general_msg_list = result.get('general_msg_list')
        for i in json.loads(general_msg_list).get('list'):
            app_msg_ext_info = i.get('app_msg_ext_info')
            title = app_msg_ext_info.get('title')
            content_url = app_msg_ext_info.get('content_url')
            author = app_msg_ext_info.get('author')
            comm_msg_info = i.get('comm_msg_info')
            datetime = comm_msg_info.get('datetime')
            publish_time = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(datetime))
            try:
                wb = load_workbook(''.join([author, '.xlsx']))
            except FileNotFoundError:
                wb = Workbook()
            ws = wb.active
            ws.append([publish_time, title, content_url])
            wb.save(filename=''.join([author, '.xlsx']))
